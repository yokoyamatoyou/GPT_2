import os
import json
import datetime
import threading
import logging
import queue
import re
import shutil

import customtkinter as ctk
from tkinter import filedialog, messagebox
import tkinter
from PIL import Image


def get_font_family(preferred: str = "Meiryo") -> str:
    """Return preferred font if available else fall back to a standard family."""
    try:
        root = tkinter.Tk()
        root.withdraw()
        families = set(root.tk.call("font", "families"))
        root.destroy()
        if preferred in families:
            return preferred
    except tkinter.TclError:
        pass
    return "Helvetica"
from openai import OpenAI
import docx
import PyPDF2
import openpyxl
import base64
from dotenv import load_dotenv
from src.agent import ReActAgent, ToTAgent, PresentationAgent
from src.main import create_evaluator
from src.memory import ConversationMemory
from src.tools import (
    get_web_scraper,
    get_sqlite_tool,
    get_graphviz_tool,
    get_mermaid_tool,
)
from src.tools.graphviz_tool import create_graphviz_diagram
from src.tools.mermaid_tool import create_mermaid_diagram

# Load environment variables from .env if present
load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

# CustomTkinterの設定
ctk.set_appearance_mode("light")
GOOGLE_THEME = os.path.join(os.path.dirname(__file__), "resources", "google.json")
ctk.set_default_color_theme(GOOGLE_THEME)

ICON_PATH = os.path.join(os.path.dirname(__file__), "resources", "app_icon.xbm")

FONT_FAMILY = get_font_family()

class ChatGPTClient:
    def __init__(self):
        """Initialize the main window and OpenAI client."""
        self.window = ctk.CTk()
        self.window.title("ChatGPT Desktop")
        self.window.geometry("1200x800")
        try:
            self.window.iconbitmap("@" + ICON_PATH)
        except Exception:
            logging.warning("Failed to set window icon")

        # モデルの初期値を環境変数から読み込む
        default_model = os.getenv("OPENAI_MODEL", "gpt-3.5-turbo")
        self.model_var = ctk.StringVar(value=default_model)
        
        # OpenAI クライアントの初期化
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            messagebox.showerror("エラー", "環境変数 OPENAI_API_KEY が設定されていません")
            logging.error("OPENAI_API_KEY is not set")
            self.window.destroy()
            return

        logging.info("Loaded OpenAI API key from environment")
        
        self.client = OpenAI(api_key=api_key)

        # 会話履歴
        self.messages = []
        self.current_title = None
        self.memory = ConversationMemory()
        self.uploaded_files = []
        self.response_queue = queue.Queue()
        self.assistant_start = None
        self._diagram_path: str | None = None
        self.agent_var = ctk.StringVar(value="chatgpt")
        self.agent_tools = [
            get_web_scraper(),
            get_sqlite_tool(),
            get_graphviz_tool(),
            get_mermaid_tool(),
        ]
        self.tools = [
            {
                "type": "function",
                "function": {
                    "name": "create_graphviz_diagram",
                    "description": "DOT言語から図を生成する。フローチャート等に適している。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "code": {"type": "string", "description": "DOT言語のコード"}
                        },
                        "required": ["code"],
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "create_mermaid_diagram",
                    "description": "Mermaid markdown-like codeから図を生成する。シーケンス図、ガントチャート等に適している。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "code": {"type": "string", "description": "Mermaid記法のコード"}
                        },
                        "required": ["code"],
                    },
                },
            },
        ]
        self.tool_funcs = {
            "create_graphviz_diagram": create_graphviz_diagram,
            "create_mermaid_diagram": create_mermaid_diagram,
        }
        
        # UI要素の作成
        self.setup_ui()
        # キュー監視処理を開始
        self.window.after(100, self.process_queue)
        
    def setup_ui(self):
        """Build all widgets and configure layout."""
        # メインコンテナ
        main_container = ctk.CTkFrame(self.window, fg_color="transparent")
        main_container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # 左側パネル（設定）
        left_panel = ctk.CTkFrame(
            main_container,
            width=300,
            fg_color="#F1F3F4",
            corner_radius=8,
            border_width=0,
        )
        left_panel.pack(side="left", fill="y", padx=(0, 10))
        left_panel.pack_propagate(False)
        
        # 設定タイトル
        settings_label = ctk.CTkLabel(left_panel, text="設定",
                                     font=(FONT_FAMILY, 22, "bold"))
        settings_label.pack(pady=20)
        
        # モデル選択
        model_label = ctk.CTkLabel(left_panel, text="モデル",
                                  font=(FONT_FAMILY, 16))
        model_label.pack(pady=(20, 5))
        
        model_menu = ctk.CTkOptionMenu(
            left_panel,
            values=[
                "gpt-3.5-turbo",
                "gpt-4.1-mini-2025-04-14",
                "gpt-4.1-nano-2025-04-14",
                "gpt-4.1-2025-04-14",
            ],
            variable=self.model_var,
            width=250,
        )
        model_menu.pack(pady=(0, 20))
        
        # 温度設定
        temp_label = ctk.CTkLabel(left_panel, text="Temperature: 0.7",
                                 font=(FONT_FAMILY, 16))
        temp_label.pack(pady=(20, 5))
        
        self.temp_slider = ctk.CTkSlider(left_panel, from_=0, to=2, number_of_steps=20,
                                        command=lambda v: temp_label.configure(text=f"Temperature: {v:.1f}"))
        self.temp_slider.set(0.7)
        self.temp_slider.pack(pady=(0, 20))

        agent_label = ctk.CTkLabel(left_panel, text="エージェント",
                                   font=(FONT_FAMILY, 16))
        agent_label.pack(pady=(10, 5))

        agent_menu = ctk.CTkOptionMenu(
            left_panel,
            values=["chatgpt", "react", "tot", "プレゼンテーション"],
            variable=self.agent_var,
            width=250,
        )
        agent_menu.pack(pady=(0, 20))
        
        # ファイルアップロードボタン
        upload_btn = ctk.CTkButton(left_panel, text="ファイルをアップロード",
                                  command=self.upload_file,
                                  font=(FONT_FAMILY, 16))
        upload_btn.pack(pady=10)
        
        # アップロードされたファイルリスト
        self.file_list_label = ctk.CTkLabel(left_panel, text="アップロードされたファイル:",
                                           font=(FONT_FAMILY, 14))
        self.file_list_label.pack(pady=(20, 5))
        
        self.file_list_text = ctk.CTkTextbox(left_panel, height=100, width=250)
        self.file_list_text.pack(pady=(0, 20))
        
        # 新しい会話ボタン
        new_chat_btn = ctk.CTkButton(left_panel, text="新しい会話",
                                    command=self.new_chat,
                                    font=(FONT_FAMILY, 16))
        new_chat_btn.pack(pady=10)

        load_chat_btn = ctk.CTkButton(
            left_panel,
            text="会話を読み込み",
            command=self.load_chat,
            font=(FONT_FAMILY, 16),
        )
        load_chat_btn.pack(pady=10)

        save_chat_btn = ctk.CTkButton(
            left_panel,
            text="会話を保存",
            command=self.save_conversation,
            font=(FONT_FAMILY, 16),
        )
        save_chat_btn.pack(pady=10)
        
        # 右側パネル（図プレビュー）
        self.diagram_panel = ctk.CTkFrame(
            main_container,
            width=250,
            fg_color="#F8F9FA",
            corner_radius=8,
            border_width=0,
        )
        self.diagram_panel.pack(side="right", fill="y")
        self.diagram_panel.pack_propagate(False)

        self.diagram_label = ctk.CTkLabel(self.diagram_panel, text="図のプレビュー", font=(FONT_FAMILY, 16))
        self.diagram_label.pack(padx=10, pady=10)

        self.save_button = ctk.CTkButton(
            self.diagram_panel,
            text="保存",
            command=lambda: self.save_diagram(),
            font=(FONT_FAMILY, 14),
            state="disabled",
        )
        self.save_button.pack(pady=(0, 10))

        self.clear_button = ctk.CTkButton(
            self.diagram_panel,
            text="クリア",
            command=lambda: self.clear_diagram(),
            font=(FONT_FAMILY, 14),
            state="disabled",
        )
        self.clear_button.pack(pady=(0, 10))

        # 右側パネル（チャット）
        right_panel = ctk.CTkFrame(
            main_container,
            fg_color="#FFFFFF",
            corner_radius=8,
            border_width=0,
        )
        right_panel.pack(side="right", fill="both", expand=True)
        
        # チャットエリア
        self.chat_display = ctk.CTkTextbox(
            right_panel,
            font=(FONT_FAMILY, 16),
            wrap="word",
            fg_color="#FFFFFF",
        )
        self.chat_display.pack(fill="both", expand=True, padx=20, pady=(20, 10))
        self.chat_display.tag_config("user_msg", background="#E8F0FE")
        self.chat_display.tag_config("assistant_msg", background="#F1F3F4")
        
        # 入力エリア
        input_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
        input_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        self.input_field = ctk.CTkEntry(input_frame, placeholder_text="メッセージを入力...",
                                       font=(FONT_FAMILY, 16), height=40)
        self.input_field.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.input_field.bind("<Return>", lambda e: self.send_message())
        
        send_btn = ctk.CTkButton(input_frame, text="送信", width=80,
                                command=self.send_message,
                                font=(FONT_FAMILY, 16))
        send_btn.pack(side="right")
        
    def upload_file(self):
        """Prompt for a file and store its contents."""
        file_path = filedialog.askopenfilename(
            title="ファイルを選択",
            filetypes=[
                ("対応ファイル", "*.docx *.pdf *.png *.jpg *.jpeg *.xlsx"),
                ("Word", "*.docx"),
                ("PDF", "*.pdf"),
                ("画像", "*.png *.jpg *.jpeg"),
                ("Excel", "*.xlsx")
            ]
        )
        
        if file_path:
            file_name = os.path.basename(file_path)
            file_ext = os.path.splitext(file_name)[1].lower()
            
            try:
                content = self.process_file(file_path, file_ext)
                self.uploaded_files.append({
                    "name": file_name,
                    "type": file_ext,
                    "content": content
                })
                
                # ファイルリストを更新
                self.update_file_list()
                
                messagebox.showinfo("成功", f"{file_name} をアップロードしました")
                
            except Exception as e:
                messagebox.showerror("エラー", f"ファイルの読み込みに失敗しました: {str(e)}")
    
    def process_file(self, file_path: str, file_ext: str) -> str:
        """ファイルタイプに応じて内容を処理"""
        if file_ext == ".docx":
            doc = docx.Document(file_path)
            return "\n".join([paragraph.text for paragraph in doc.paragraphs])
            
        elif file_ext == ".pdf":
            with open(file_path, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    extracted_text = page.extract_text()
                    if extracted_text:
                        text += extracted_text + "\n"
                return text
                
        elif file_ext in [".png", ".jpg", ".jpeg"]:
            # 画像をbase64エンコード
            with open(file_path, "rb") as img_file:
                return base64.b64encode(img_file.read()).decode('utf-8')
                
        elif file_ext == ".xlsx":
            workbook = openpyxl.load_workbook(file_path)
            sheets_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        sheet_data.append(list(row))
                
                sheets_data[sheet_name] = sheet_data
            
            # シート情報を文字列として返す
            result = f"Excelファイル: {len(workbook.sheetnames)}個のシート\n"
            for sheet_name, data in sheets_data.items():
                result += f"\n【シート: {sheet_name}】\n"
                result += f"行数: {len(data)}\n"
                if data:
                    result += f"列数: {len(data[0])}\n"
                    # 最初の数行を表示
                    for i, row in enumerate(data[:5]):
                        result += f"行{i+1}: {row}\n"
                    if len(data) > 5:
                        result += "...\n"
            
            return result
        
        return ""
    
    def update_file_list(self):
        """アップロードされたファイルリストを更新"""
        self.file_list_text.configure(state="normal")
        self.file_list_text.delete("1.0", "end")
        for file in self.uploaded_files:
            self.file_list_text.insert("end", f"• {file['name']}\n")
        self.file_list_text.configure(state="disabled")
    
    def send_message(self):
        """Handle user input and start fetching a reply."""
        user_message = self.input_field.get().strip()
        if not user_message:
            return
        
        # メッセージをクリア
        self.input_field.delete(0, "end")
        
        # ユーザーメッセージを表示
        self.chat_display.configure(state="normal")
        start = self.chat_display.index("end") if hasattr(self.chat_display, "index") else None
        self.chat_display.insert("end", f"\n👤 You: {user_message}\n\n")
        if start is not None and hasattr(self.chat_display, "tag_add"):
            end = self.chat_display.index("end")
            self.chat_display.tag_add("user_msg", start, end)
        self.chat_display.see("end")
        self.chat_display.configure(state="disabled")
        
        # ファイル情報を含めたメッセージを作成
        # OpenAI APIは、"user"ロールのメッセージcontentに直接画像を含めることを想定
        # "vision"モデル (例: gpt-4-vision-preview, gpt-4o) は content に配列を受け付けます。
        # ここでは、ファイルの内容をテキストとして含めることを前提としています。
        # 画像ファイルがある場合、その内容 (base64) は直接メッセージに含めず、
        # 別途、GPT-4V (Vision) などのマルチモーダルモデルのAPIコール時に適切に処理する必要があります。
        # 現在のコードでは、画像はbase64エンコードされた文字列としてcontentに含めていますが、
        # これが直接的にテキストモデルで解釈されるわけではありません。
        # gpt-4oのようなマルチモーダルモデルでは、contentに画像データを含めるための特定の形式が必要です。
        # (例: `{"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_string}"}}`)
        # 今回の修正はモデル名の変更のみに留め、このロジックは変更しません。

        content_parts = [{"type": "text", "text": user_message}]

        if not self.messages:
            system_prompt = (
                "あなたは優秀なAIアシスタントです。ユーザーの質問が曖昧だと判断した場合、"
                "通常の回答の後に『【プロンプトアドバイス】』という見出しを付け、より具体的な質問の例を2〜3個提示してください。"
                "例: 『ラーメンについて教えて』→『東京でおすすめの醤油ラーメンのお店は？』"
            )
            self.messages.append({"role": "system", "content": system_prompt})
        
        if self.uploaded_files:
            file_info_text = "\n\n【アップロードされたファイル情報】\n"
            for file in self.uploaded_files:
                if file['type'] in ['.docx', '.pdf', '.xlsx']:
                    file_info_text += f"\n--- {file['name']} ---\n{file['content'][:1000]}...\n" # 長すぎる内容は省略
                elif file['type'] in ['.png', '.jpg', '.jpeg']:
                    # gpt-4oなどのマルチモーダルモデルの場合、画像は特別な形式で渡す
                    # ここでは、メッセージに追加のテキスト情報としてファイル名のみ含めるか、
                    # あるいは、content_partsに画像データを追加する処理が必要。
                    # 今回のモデル名変更のリクエストでは、この部分のロジックは変更しない。
                    # 単純にテキストとしてファイル名を付加する例：
                    file_info_text += f"\n画像ファイル: {file['name']} (内容は別途送信されます)\n"
                    # もしgpt-4oに画像を直接渡すなら、以下のような形式でcontent_partsに追加
                    # image_data = {
                    #     "type": "image_url",
                    #     "image_url": {
                    #         "url": f"data:image/{file['type'][1:]};base64,{file['content']}"
                    #     }
                    # }
                    # content_parts.append(image_data)

            # ユーザーメッセージのテキストパートにファイル情報を追加
            content_parts[0]["text"] += file_info_text
        
        # メッセージを履歴に追加
        self.messages.append({"role": "user", "content": content_parts})
        
        # 初回メッセージの場合、タイトルを生成
        user_count = sum(1 for m in self.messages if m.get("role") == "user")
        if user_count == 1:
            self.generate_title(user_message)
        
        # エージェント種別に応じて応答を取得
        if getattr(self, "agent_var", None) and self.agent_var.get() != "chatgpt":
            agent_type = self.agent_var.get()
            threading.Thread(target=self.run_agent, args=(agent_type, user_message), daemon=True).start()
        else:
            threading.Thread(target=self.get_response, daemon=True).start()
    
    def get_response(self):
        """Stream the assistant's reply, execute tool calls, and push updates."""
        try:
            self.response_queue.put("🤖 Assistant: ")
            while True:
                response_text = ""
                tool_data: dict[str, dict[str, str]] = {}

                params = {
                    "model": self.model_var.get(),
                    "messages": self.messages,
                    "temperature": self.temp_slider.get(),
                    "stream": True,
                }
                if getattr(self, "tools", None):
                    params["tools"] = self.tools
                    params["tool_choice"] = "auto"

                stream = self.client.chat.completions.create(**params)
                finish_reason = None

                for chunk in stream:
                    choice = chunk.choices[0]
                    delta = choice.delta
                    finish_reason = choice.finish_reason

                    if getattr(delta, "content", None) is not None:
                        content = delta.content
                        response_text += content
                        self.response_queue.put(content)

                    calls = getattr(delta, "tool_calls", None)
                    if calls:
                        for c in calls:
                            info = tool_data.setdefault(c.id, {"name": "", "args": ""})
                            if getattr(c.function, "name", None):
                                info["name"] = c.function.name
                            if getattr(c.function, "arguments", None):
                                info["args"] += c.function.arguments

                self.response_queue.put("\n")

                if tool_data and finish_reason == "tool_calls":
                    assistant_msg = {
                        "role": "assistant",
                        "content": response_text,
                        "tool_calls": [
                            {
                                "id": cid,
                                "type": "function",
                                "function": {"name": d["name"], "arguments": d["args"]},
                            }
                            for cid, d in tool_data.items()
                        ],
                    }
                    self.messages.append(assistant_msg)

                    for cid, d in tool_data.items():
                        func = self.tool_funcs.get(d["name"])
                        if func:
                            try:
                                args = json.loads(d["args"] or "{}")
                                result = func(**args)
                            except Exception as exc:
                                result = f"Tool execution failed: {exc}"
                        else:
                            result = f"Unknown tool: {d['name']}"
                        self.messages.append({"role": "tool", "tool_call_id": cid, "content": result})

                    # Continue looping to stream assistant's follow-up answer
                    continue

                # 通常の応答を保存して終了
                self.messages.append({"role": "assistant", "content": response_text})
                match = re.search(r"(/[^\s]+\.png)", response_text)
                if match and os.path.isfile(match.group(1)):
                    self.response_queue.put(f"__DIAGRAM__{match.group(1)}")
                self.response_queue.put("__SAVE__")
                break

        except Exception as e:
            self.response_queue.put(f"\n\nエラー: {str(e)}\n")

    def simple_llm(self, prompt: str) -> str:
        """Call the OpenAI API synchronously and return the message text."""
        resp = self.client.chat.completions.create(
            model=self.model_var.get(),
            messages=[{"role": "user", "content": prompt}],
            temperature=self.temp_slider.get(),
        )
        return resp.choices[0].message.content

    def run_agent(self, agent_type: str, question: str) -> None:
        """Execute the selected agent and stream steps to the queue."""
        try:
            self.response_queue.put("🤖 Assistant: ")
            if agent_type == "react":
                agent = ReActAgent(self.simple_llm, self.agent_tools, self.memory)
            elif agent_type == "tot":
                evaluator = create_evaluator(self.simple_llm)
                agent = ToTAgent(self.simple_llm, evaluator, memory=self.memory)
            elif agent_type == "プレゼンテーション":
                agent = PresentationAgent(self.simple_llm)
            else:
                self.response_queue.put("未対応のエージェントです\n")
                return
            response_text = ""
            for step in agent.run_iter(question):
                response_text += step + "\n"
                self.response_queue.put(step + "\n")
            self.messages.append({"role": "user", "content": question})
            self.messages.append({"role": "assistant", "content": response_text})
            match = re.search(r"(/[^\s]+\.png)", response_text)
            if match and os.path.isfile(match.group(1)):
                self.response_queue.put(f"__DIAGRAM__{match.group(1)}")
            self.response_queue.put("__SAVE__")
        except Exception as exc:
            self.response_queue.put(f"\n\nエラー: {exc}\n")
    
    def generate_title(self, first_message: str):
        """最初のメッセージからタイトルを生成"""
        try:
            # タイトル生成はシンプルなモデルで十分
            response = self.client.chat.completions.create(
                model="gpt-3.5-turbo", # タイトル生成用モデルは変更なし
                messages=[
                    {"role": "system", "content": "ユーザーのメッセージから、短く簡潔な会話のタイトルを生成してください。20文字以内で。"},
                    {"role": "user", "content": first_message}
                ],
                temperature=0.5,
                max_tokens=50
            )
            
            self.current_title = response.choices[0].message.content.strip()
            self.window.title(f"ChatGPT Desktop - {self.current_title}")
            
        except Exception:
            logging.exception("Failed to generate title")
            self.current_title = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            self.window.title(f"ChatGPT Desktop - {self.current_title}")  # エラー時もタイトル設定
    
    def save_conversation(self, show_popup: bool = True):
        """会話をJSONファイルとして保存."""
        if not self.current_title:
            return
        
        # conversationsディレクトリがなければ作成
        if not os.path.exists("conversations"):
            os.makedirs("conversations")
            
        filename_base = f"{self.current_title}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
        # ファイル名に使えない文字を置換
        filename_safe = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in filename_base)
        filename = os.path.join("conversations", f"{filename_safe}.json")
        
        # uploaded_filesのcontentは保存しない (大きすぎる可能性があるため)
        files_metadata = []
        for f_info in self.uploaded_files:
            files_metadata.append({
                "name": f_info["name"],
                "type": f_info["type"]
            })

        conversation_data = {
            "title": self.current_title,
            "timestamp": datetime.datetime.now().isoformat(),
            "model": self.model_var.get(),
            "messages": self.messages, # メッセージ履歴をそのまま保存
            "uploaded_files_metadata": files_metadata # contentは含めない
        }
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(conversation_data, f, ensure_ascii=False, indent=2)
            if show_popup:
                try:
                    messagebox.showinfo("保存完了", f"会話を {filename} に保存しました")
                except tkinter.TclError:
                    pass
        except Exception as e:
            if show_popup:
                messagebox.showerror("保存エラー", f"会話の保存に失敗しました: {str(e)}")
            else:
                logging.error("会話の保存に失敗しました: %s", e)

    
    def new_chat(self):
        """新しい会話を開始"""
        self.messages = []
        self.current_title = None
        self.uploaded_files = []
        try:
            if isinstance(self.memory, ConversationMemory):
                self.memory = ConversationMemory()
            elif hasattr(self.memory, "messages"):
                self.memory.messages.clear()
        except Exception:
            logging.warning("Failed to reset memory", exc_info=True)

        self.chat_display.configure(state="normal")
        self.chat_display.delete("1.0", "end")
        self.chat_display.insert("1.0", "新しい会話を開始しました。\n")
        self.chat_display.configure(state="disabled")
        
        self.file_list_text.configure(state="normal")
        self.file_list_text.delete("1.0", "end")
        self.file_list_text.configure(state="disabled")
        
        self.window.title("ChatGPT Desktop")

    def load_chat(self):
        """Open a saved conversation file and load its content."""
        file_path = filedialog.askopenfilename(
            title="会話を選択",
            filetypes=[("Conversation", "*.json")],
        )
        if file_path:
            self.load_conversation(file_path)

    def load_conversation(self, file_path: str):
        """Load conversation from a JSON file created by save_conversation."""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("読み込みエラー", f"会話の読み込みに失敗しました: {str(e)}")
            return

        self.current_title = data.get("title")
        if self.current_title:
            self.window.title(f"ChatGPT Desktop - {self.current_title}")
        self.messages = data.get("messages", [])
        meta = data.get("uploaded_files_metadata", [])
        self.uploaded_files = [{"name": m["name"], "type": m["type"]} for m in meta]

        # Refresh displays
        self.update_file_list()
        self.chat_display.configure(state="normal")
        self.chat_display.delete("1.0", "end")
        for msg in self.messages:
            role = msg.get("role")
            content = msg.get("content", "")
            if isinstance(content, list):
                # content may be structured as list of parts
                content = "".join(part.get("text", "") for part in content)
            prefix = "👤 You" if role == "user" else "🤖 Assistant"
            start = self.chat_display.index("end") if hasattr(self.chat_display, "index") else None
            self.chat_display.insert("end", f"\n{prefix}: {content}\n\n")
            if start is not None and hasattr(self.chat_display, "tag_add"):
                end = self.chat_display.index("end")
                tag = "user_msg" if role == "user" else "assistant_msg"
                self.chat_display.tag_add(tag, start, end)
        self.chat_display.configure(state="disabled")

    def display_diagram(self, path: str) -> None:
        """Preview a diagram PNG and enable saving."""
        try:
            img = Image.open(path)
            preview = ctk.CTkImage(light_image=img, size=(200, 200))
        except Exception:
            logging.exception("Failed to load diagram %s", path)
            return
        self.diagram_label.configure(image=preview, text="")
        self.diagram_label.image = preview
        self.save_button.configure(state="normal")
        self.clear_button.configure(state="normal")
        self._diagram_path = path

    def save_diagram(self) -> None:
        """Save the currently previewed diagram to a location chosen by the user."""
        if not getattr(self, "_diagram_path", None):
            return
        dest = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG", "*.png")])
        if dest:
            try:
                shutil.copy(self._diagram_path, dest)
            except Exception as exc:
                messagebox.showerror("保存エラー", str(exc))
            else:
                messagebox.showinfo("保存完了", f"図を {dest} に保存しました")

    def clear_diagram(self) -> None:
        """Remove the current diagram preview and disable related buttons."""
        self.diagram_label.configure(image=None, text="図のプレビュー")
        self.diagram_label.image = None
        self.save_button.configure(state="disabled")
        self.clear_button.configure(state="disabled")
        self._diagram_path = None

    def process_queue(self):
        """キューからのメッセージをGUIに反映"""
        try:
            while True:
                item = self.response_queue.get_nowait()
                if item.startswith("__DIAGRAM__"):
                    self.display_diagram(item[len("__DIAGRAM__"):])
                    continue
                if item == "__SAVE__":
                    threading.Thread(
                        target=self.save_conversation,
                        kwargs={"show_popup": False},
                        daemon=True,
                    ).start()
                    continue
                self.chat_display.configure(state="normal")
                if item.startswith("🤖 Assistant: "):
                    self.assistant_start = self.chat_display.index("end") if hasattr(self.chat_display, "index") else None
                    self.chat_display.insert("end", item)
                else:
                    self.chat_display.insert("end", item)
                if self.assistant_start is not None and item.endswith("\n") and hasattr(self.chat_display, "tag_add"):
                    end = self.chat_display.index("end")
                    self.chat_display.tag_add("assistant_msg", self.assistant_start, end)
                    self.assistant_start = None
                self.chat_display.see("end")
                self.chat_display.configure(state="disabled")
        except queue.Empty:
            pass
        self.window.after(100, self.process_queue)
    
    def run(self):
        """Start the application event loop."""
        # 初期化時にチャット表示とファイルリストをdisabledに
        self.chat_display.configure(state="disabled")
        self.file_list_text.configure(state="disabled")
        self.window.mainloop()

if __name__ == "__main__":
    app = ChatGPTClient()
    app.run()
