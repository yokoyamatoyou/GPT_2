import os
import json
import datetime
import threading
import logging

import customtkinter as ctk
from tkinter import filedialog, messagebox
from openai import OpenAI
import docx
import PyPDF2
from PIL import Image
import openpyxl
import base64
import io
from typing import Optional, List, Dict
from dotenv import load_dotenv

# Load environment variables from .env if present
load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

# CustomTkinterの設定
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

class ChatGPTClient:
    def __init__(self):
        self.window = ctk.CTk()
        self.window.title("ChatGPT Desktop")
        self.window.geometry("1200x800")
        
        # OpenAI クライアントの初期化
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            messagebox.showerror("エラー", "環境変数 OPENAI_API_KEY が設定されていません")
            logging.error("OPENAI_API_KEY is not set")
            self.window.destroy()
            return

        logging.info("Loaded OpenAI API key from environment")
        
        self.client = OpenAI(api_key=api_key)
        
        # 会話履歴
        self.messages = []
        self.current_title = None
        self.uploaded_files = []
        
        # UI要素の作成
        self.setup_ui()
        
    def setup_ui(self):
        # メインコンテナ
        main_container = ctk.CTkFrame(self.window, fg_color="transparent")
        main_container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # 左側パネル（設定）
        left_panel = ctk.CTkFrame(main_container, width=300, fg_color="#f0f0f0")
        left_panel.pack(side="left", fill="y", padx=(0, 10))
        left_panel.pack_propagate(False)
        
        # 設定タイトル
        settings_label = ctk.CTkLabel(left_panel, text="設定", 
                                     font=("SF Pro Display", 20, "bold"))
        settings_label.pack(pady=20)
        
        # モデル選択
        model_label = ctk.CTkLabel(left_panel, text="モデル", 
                                  font=("SF Pro Display", 14))
        model_label.pack(pady=(20, 5))
        
        self.model_var = ctk.StringVar(value="gpt-4.1-mini-2025-04-14") # <--- 変更点: デフォルトモデル
        model_menu = ctk.CTkOptionMenu(left_panel, 
                                      values=["gpt-4.1-mini-2025-04-14", 
                                              "gpt-4.1-nano-2025-04-14", 
                                              "gpt-4.1-2025-04-14"], # <--- 変更点: モデルリスト
                                      variable=self.model_var, width=250)
        model_menu.pack(pady=(0, 20))
        
        # 温度設定
        temp_label = ctk.CTkLabel(left_panel, text="Temperature: 0.7", 
                                 font=("SF Pro Display", 14))
        temp_label.pack(pady=(20, 5))
        
        self.temp_slider = ctk.CTkSlider(left_panel, from_=0, to=2, number_of_steps=20,
                                        command=lambda v: temp_label.configure(text=f"Temperature: {v:.1f}"))
        self.temp_slider.set(0.7)
        self.temp_slider.pack(pady=(0, 20))
        
        # ファイルアップロードボタン
        upload_btn = ctk.CTkButton(left_panel, text="ファイルをアップロード", 
                                  command=self.upload_file,
                                  font=("SF Pro Display", 14))
        upload_btn.pack(pady=10)
        
        # アップロードされたファイルリスト
        self.file_list_label = ctk.CTkLabel(left_panel, text="アップロードされたファイル:", 
                                           font=("SF Pro Display", 12))
        self.file_list_label.pack(pady=(20, 5))
        
        self.file_list_text = ctk.CTkTextbox(left_panel, height=100, width=250)
        self.file_list_text.pack(pady=(0, 20))
        
        # 新しい会話ボタン
        new_chat_btn = ctk.CTkButton(left_panel, text="新しい会話", 
                                    command=self.new_chat,
                                    font=("SF Pro Display", 14))
        new_chat_btn.pack(pady=10)
        
        # 右側パネル（チャット）
        right_panel = ctk.CTkFrame(main_container, fg_color="#ffffff")
        right_panel.pack(side="right", fill="both", expand=True)
        
        # チャットエリア
        self.chat_display = ctk.CTkTextbox(right_panel, font=("SF Pro Text", 14),
                                          wrap="word", fg_color="#f8f8f8")
        self.chat_display.pack(fill="both", expand=True, padx=20, pady=(20, 10))
        
        # 入力エリア
        input_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
        input_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        self.input_field = ctk.CTkEntry(input_frame, placeholder_text="メッセージを入力...",
                                       font=("SF Pro Text", 14), height=40)
        self.input_field.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.input_field.bind("<Return>", lambda e: self.send_message())
        
        send_btn = ctk.CTkButton(input_frame, text="送信", width=80,
                                command=self.send_message,
                                font=("SF Pro Display", 14))
        send_btn.pack(side="right")
        
    def upload_file(self):
        file_path = filedialog.askopenfilename(
            title="ファイルを選択",
            filetypes=[
                ("対応ファイル", "*.docx *.pdf *.png *.jpg *.jpeg *.xlsx"),
                ("Word", "*.docx"),
                ("PDF", "*.pdf"),
                ("画像", "*.png *.jpg *.jpeg"),
                ("Excel", "*.xlsx")
            ]
        )
        
        if file_path:
            file_name = os.path.basename(file_path)
            file_ext = os.path.splitext(file_name)[1].lower()
            
            try:
                content = self.process_file(file_path, file_ext)
                self.uploaded_files.append({
                    "name": file_name,
                    "type": file_ext,
                    "content": content
                })
                
                # ファイルリストを更新
                self.update_file_list()
                
                messagebox.showinfo("成功", f"{file_name} をアップロードしました")
                
            except Exception as e:
                messagebox.showerror("エラー", f"ファイルの読み込みに失敗しました: {str(e)}")
    
    def process_file(self, file_path: str, file_ext: str) -> str:
        """ファイルタイプに応じて内容を処理"""
        if file_ext == ".docx":
            doc = docx.Document(file_path)
            return "\n".join([paragraph.text for paragraph in doc.paragraphs])
            
        elif file_ext == ".pdf":
            with open(file_path, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    extracted_text = page.extract_text()
                    if extracted_text:
                        text += extracted_text + "\n"
                return text
                
        elif file_ext in [".png", ".jpg", ".jpeg"]:
            # 画像をbase64エンコード
            with open(file_path, "rb") as img_file:
                return base64.b64encode(img_file.read()).decode('utf-8')
                
        elif file_ext == ".xlsx":
            workbook = openpyxl.load_workbook(file_path)
            sheets_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        sheet_data.append(list(row))
                
                sheets_data[sheet_name] = sheet_data
            
            # シート情報を文字列として返す
            result = f"Excelファイル: {len(workbook.sheetnames)}個のシート\n"
            for sheet_name, data in sheets_data.items():
                result += f"\n【シート: {sheet_name}】\n"
                result += f"行数: {len(data)}\n"
                if data:
                    result += f"列数: {len(data[0])}\n"
                    # 最初の数行を表示
                    for i, row in enumerate(data[:5]):
                        result += f"行{i+1}: {row}\n"
                    if len(data) > 5:
                        result += "...\n"
            
            return result
        
        return ""
    
    def update_file_list(self):
        """アップロードされたファイルリストを更新"""
        self.file_list_text.configure(state="normal")
        self.file_list_text.delete("1.0", "end")
        for file in self.uploaded_files:
            self.file_list_text.insert("end", f"• {file['name']}\n")
        self.file_list_text.configure(state="disabled")
    
    def send_message(self):
        user_message = self.input_field.get().strip()
        if not user_message:
            return
        
        # メッセージをクリア
        self.input_field.delete(0, "end")
        
        # ユーザーメッセージを表示
        self.chat_display.configure(state="normal")
        self.chat_display.insert("end", f"\n👤 You: {user_message}\n\n")
        self.chat_display.see("end")
        self.chat_display.configure(state="disabled")
        
        # ファイル情報を含めたメッセージを作成
        # OpenAI APIは、"user"ロールのメッセージcontentに直接画像を含めることを想定
        # "vision"モデル (例: gpt-4-vision-preview, gpt-4o) は content に配列を受け付けます。
        # ここでは、ファイルの内容をテキストとして含めることを前提としています。
        # 画像ファイルがある場合、その内容 (base64) は直接メッセージに含めず、
        # 別途、GPT-4V (Vision) などのマルチモーダルモデルのAPIコール時に適切に処理する必要があります。
        # 現在のコードでは、画像はbase64エンコードされた文字列としてcontentに含めていますが、
        # これが直接的にテキストモデルで解釈されるわけではありません。
        # gpt-4oのようなマルチモーダルモデルでは、contentに画像データを含めるための特定の形式が必要です。
        # (例: `{"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_string}"}}`)
        # 今回の修正はモデル名の変更のみに留め、このロジックは変更しません。

        content_parts = [{"type": "text", "text": user_message}]
        
        if self.uploaded_files:
            file_info_text = "\n\n【アップロードされたファイル情報】\n"
            for file in self.uploaded_files:
                if file['type'] in ['.docx', '.pdf', '.xlsx']:
                    file_info_text += f"\n--- {file['name']} ---\n{file['content'][:1000]}...\n" # 長すぎる内容は省略
                elif file['type'] in ['.png', '.jpg', '.jpeg']:
                    # gpt-4oなどのマルチモーダルモデルの場合、画像は特別な形式で渡す
                    # ここでは、メッセージに追加のテキスト情報としてファイル名のみ含めるか、
                    # あるいは、content_partsに画像データを追加する処理が必要。
                    # 今回のモデル名変更のリクエストでは、この部分のロジックは変更しない。
                    # 単純にテキストとしてファイル名を付加する例：
                    file_info_text += f"\n画像ファイル: {file['name']} (内容は別途送信されます)\n"
                    # もしgpt-4oに画像を直接渡すなら、以下のような形式でcontent_partsに追加
                    # image_data = {
                    #     "type": "image_url",
                    #     "image_url": {
                    #         "url": f"data:image/{file['type'][1:]};base64,{file['content']}"
                    #     }
                    # }
                    # content_parts.append(image_data)

            # ユーザーメッセージのテキストパートにファイル情報を追加
            content_parts[0]["text"] += file_info_text
        
        # メッセージを履歴に追加
        self.messages.append({"role": "user", "content": content_parts})
        
        # 初回メッセージの場合、タイトルを生成
        if len(self.messages) == 1:
            self.generate_title(user_message)
        
        # 別スレッドで応答を取得
        threading.Thread(target=self.get_response, daemon=True).start()
    
    def get_response(self):
        try:
            self.chat_display.configure(state="normal")
            self.chat_display.insert("end", "🤖 Assistant: ")
            
            response_text = ""
            # gpt-4oなどマルチモーダルモデルを正しく使う場合、入力メッセージの形式に注意
            # self.messages は `[{"role": "user", "content": [{"type": "text", ...}, {"type": "image_url", ...}]}]` の形式
            
            # 現在のself.messagesのcontentは、send_messageで構築されたcontent_partsを想定
            messages_for_api = self.messages 

            stream = self.client.chat.completions.create(
                model=self.model_var.get(),
                messages=messages_for_api,
                temperature=self.temp_slider.get(),
                stream=True
            )
            
            for chunk in stream:
                if chunk.choices[0].delta.content is not None:
                    content = chunk.choices[0].delta.content
                    response_text += content
                    self.chat_display.insert("end", content)
                    self.chat_display.see("end")
                    self.window.update_idletasks() # update()から変更
            
            self.chat_display.insert("end", "\n")
            self.chat_display.configure(state="disabled")
            
            # アシスタントの応答を履歴に追加
            self.messages.append({"role": "assistant", "content": response_text})
            
            # 会話を保存
            self.save_conversation()
            
        except Exception as e:
            self.chat_display.configure(state="normal")
            self.chat_display.insert("end", f"\n\nエラー: {str(e)}\n")
            self.chat_display.see("end")
            self.chat_display.configure(state="disabled")
    
    def generate_title(self, first_message: str):
        """最初のメッセージからタイトルを生成"""
        try:
            # タイトル生成はシンプルなモデルで十分
            response = self.client.chat.completions.create(
                model="gpt-3.5-turbo", # タイトル生成用モデルは変更なし
                messages=[
                    {"role": "system", "content": "ユーザーのメッセージから、短く簡潔な会話のタイトルを生成してください。20文字以内で。"},
                    {"role": "user", "content": first_message}
                ],
                temperature=0.5,
                max_tokens=50
            )
            
            self.current_title = response.choices[0].message.content.strip()
            self.window.title(f"ChatGPT Desktop - {self.current_title}")
            
        except Exception: # エラーハンドリングを少し具体的に
            self.current_title = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            self.window.title(f"ChatGPT Desktop - {self.current_title}") # エラー時もタイトル設定
    
    def save_conversation(self):
        """会話をJSONファイルとして保存"""
        if not self.current_title:
            return
        
        # conversationsディレクトリがなければ作成
        if not os.path.exists("conversations"):
            os.makedirs("conversations")
            
        filename_base = f"{self.current_title}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
        # ファイル名に使えない文字を置換
        filename_safe = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in filename_base)
        filename = os.path.join("conversations", f"{filename_safe}.json")
        
        # uploaded_filesのcontentは保存しない (大きすぎる可能性があるため)
        files_metadata = []
        for f_info in self.uploaded_files:
            files_metadata.append({
                "name": f_info["name"],
                "type": f_info["type"]
            })

        conversation_data = {
            "title": self.current_title,
            "timestamp": datetime.datetime.now().isoformat(),
            "model": self.model_var.get(),
            "messages": self.messages, # メッセージ履歴をそのまま保存
            "uploaded_files_metadata": files_metadata # contentは含めない
        }
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(conversation_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("保存エラー", f"会話の保存に失敗しました: {str(e)}")

    
    def new_chat(self):
        """新しい会話を開始"""
        self.messages = []
        self.current_title = None
        self.uploaded_files = []
        
        self.chat_display.configure(state="normal")
        self.chat_display.delete("1.0", "end")
        self.chat_display.insert("1.0", "新しい会話を開始しました。\n")
        self.chat_display.configure(state="disabled")
        
        self.file_list_text.configure(state="normal")
        self.file_list_text.delete("1.0", "end")
        self.file_list_text.configure(state="disabled")
        
        self.window.title("ChatGPT Desktop")
    
    def run(self):
        # 初期化時にチャット表示とファイルリストをdisabledに
        self.chat_display.configure(state="disabled")
        self.file_list_text.configure(state="disabled")
        self.window.mainloop()

if __name__ == "__main__":
    app = ChatGPTClient()
    app.run()